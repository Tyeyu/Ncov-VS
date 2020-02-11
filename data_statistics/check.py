#地市新增核对表格核对
import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.styles import  PatternFill
#读取文件
wb = openpyxl.load_workbook('地市新增核对表格2020-2253.xlsx')
sheet=wb['省级卫健委新增确认']

time_dict={}
postion_dict={}
country_dict={}
for e_row in range(2,sheet.max_row+1):
    #print(e_row)
    a=sheet.cell(row=e_row,column=1).value
    b=sheet.cell(row=e_row,column=2).value
    c=sheet.cell(row=e_row,column=3).value
    d=sheet.cell(row=e_row,column=4).value
    e=sheet.cell(row=e_row,column=5).value
    f=sheet.cell(row=e_row,column=6).value
    g=sheet.cell(row=e_row,column=7).value
    h=sheet.cell(row=e_row,column=8).value

    if(b==None):
        continue
    if(h==None):
        h=0
    
    if((b=='国家级' or b=='省级')and(c!='香港' and c!='澳门' and c!='台湾')):
        e_time = str(a.month) + "月" + str(a.day) + "日"
        if(e_time not in country_dict):
            country_dict[e_time]={}
            country_dict[e_time]['省级总和']=[0,0,0]
        if(b=='国家级'):
            country_dict[e_time]['国家级']=[0,0,0]
            if (e!=None):
                country_dict[e_time]['国家级'][0]=e+h
            if(f!=None):
                country_dict[e_time]['国家级'][1]=f
            if(g!=None):
                country_dict[e_time]['国家级'][2]=g
        elif(b=='省级'):
            country_dict[e_time][c]=[0,0,0]
            if(e!=None):
                country_dict[e_time][c][0]=e+h
            if(f!=None):
                country_dict[e_time][c][1]=f
            if(g!=None):
                country_dict[e_time][c][2]=g

            if(e!=None):
                country_dict[e_time]['省级总和'][0]+=(e+h)
            if(f!=None):
                country_dict[e_time]['省级总和'][1]+=f
            if(g!=None):
                country_dict[e_time]['省级总和'][2]+=g
            
    if((b=='地区级' or b=='省级')and(c!='香港' and c!='澳门' and c!='台湾')):
        e_time = str(a.month) + "月" + str(a.day) + "日"
        if(e_time not in time_dict):
            time_dict[e_time]={}
        if(c not in time_dict[e_time]):
            time_dict[e_time][c]=[0, 0, 0, 0, 0, 0]
        if(b=='地区级'):
            if(e!=None):
                time_dict[e_time][c][0]+=e
            if(f!=None):
                time_dict[e_time][c][1]+=f
            if(g!=None):
                time_dict[e_time][c][2]+=g
        elif(b=='省级'):
            if(e!=None):
                time_dict[e_time][c][3]=e
            if(f!=None):
                time_dict[e_time][c][4]=f
            if(g!=None):
                time_dict[e_time][c][5]=g
        
        if(c not in postion_dict):
            postion_dict[c]=[0,0,0,0,0,0,0] #后一位是省的核减和
        if(b=='地区级'):
            if(e!=None):
                postion_dict[c][0]+=e #总新增
            if(f!=None):
                postion_dict[c][1]+=f
            if(g!=None):
                postion_dict[c][2]+=g
        elif(b=='省级'):             #总新增
            if(e!=None):
                postion_dict[c][3]+=e
            if(f!=None):
                postion_dict[c][4]+=f
            if(g!=None):
                postion_dict[c][5]+=g
            postion_dict[c][6]+=h

# print(country_dict)
# print(time_dict)
# print(postion_dict)
f1 = open('test.txt','w')

f1.write("------------------省级复核，每天-----------------\n")
time_keys=list(time_dict.keys())
for i in range(0,len(time_dict)):
    ShengFen=list(time_dict[time_keys[i]].keys())
    for j in range(0,len(time_dict[time_keys[i]])):
        a=time_dict[time_keys[i]][ShengFen[j]]
        if ShengFen[j]==None:
                ShengFen[j]='无数据'
        if a[0]!=a[3]:
                error_inf=ShengFen[j]+','+time_keys[i]+',地区新增确诊病例累加值为'+str(int(a[0]))+',省级新增'+str(int(a[3]))
                f1.write(error_inf)
                f1.write('\n')
        if a[1]!=a[4]:
                error_inf = ShengFen[j] + ',' + time_keys[i] + ',地区新增治愈出院数累加值为' + str(int(a[1])) + ',省级新增' + str(int(a[4]))
                f1.write(error_inf)
                f1.write('\n')
        if a[2] != a[5]:
                error_inf = ShengFen[j] + ',' + time_keys[i] + ',地区新增死亡数累加值为' + str(int(a[2])) + ',省级新增' + str(int(a[5]))
                f1.write(error_inf)
                f1.write('\n')
f1.write("------------------省级复核，累积-----------------\n")
postion_keys=list(postion_dict.keys())
for i in range(0,len(postion_dict)):
    a=postion_dict[postion_keys[i]]
    # if postion_keys[i] == None:
    #     postion_keys[i]='无数据'
    # if a[0]!=a[3]:
    #     error_inf=postion_keys[i] + ',各地区新增确诊病例累积'+str(int(a[0]))+',省级新增'+str(int(a[3]))
    #     print(error_inf)
    # if a[1]!=a[4]:
    #     error_inf = postion_keys[i] + ',各地区新增治愈出院数累积' + str(int(a[1])) + ',省级新增' + str(int(a[4]))
    #     print(error_inf)
    # if a[2] != a[5]:
    #     error_inf = postion_keys[i] + ',各地区新增死亡数累积' + str(int(a[2])) + ',省级新增' + str(int(a[5]))
    #     print(error_inf)
    ans=postion_keys[i]+',各地区新增确诊病例累积'+str(int(a[0]))+',省级新增'+str(int(a[3]))+',省级累计'+str(int(a[3]+a[6]))+'(核减'+str(int(a[6]))+')'
    f1.write(ans)
    f1.write('\n')
f1.write("------------------全国级复核-----------------\n")
country_keys=list(country_dict.keys())
for i in range(0,len(country_dict)):
    a=country_dict[country_keys[i]]
    #print(country_keys[i])
    if '国家级' not in list(a.keys()):
        error_inf = country_keys[i] + '没有统计国家数据'
        f1.write(error_inf)
        f1.write('\n')
        continue
    if a['国家级'][0]!=a['省级总和'][0]:
        error_inf = country_keys[i] + ',各省新增确诊病例累积' + str(int(a['省级总和'][0])) + ',国家级新增' + str(int(a['国家级'][0]))
        f1.write(error_inf)
    if a['国家级'][1]!=a['省级总和'][1]:
        error_inf = country_keys[i] + ',各省新增治愈出院数累积' + str(int(a['省级总和'][1])) + ',国家级新增' + str(int(a['国家级'][1]))
        f1.write(error_inf)
        f1.write('\n')
    if a['国家级'][2]!=a['省级总和'][2]:
        error_inf = country_keys[i] + ',各省新增死亡数累积' + str(int(a['省级总和'][2])) + ',国家级新增' + str(int(a['国家级'][2]))
        f1.write(error_inf)
        f1.write('\n')
f1.close()
print('done')