#地市新增核对表格核对
import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.styles import  PatternFill
#读取文件
wb = openpyxl.load_workbook('地市新增核对表格.xlsx')
sheet=wb['工作表1']

time_list={}
postion_list={}
country_list={}
for e_row in range(2,sheet.max_row+1):
    #print(e_row)
    if sheet.cell(row=e_row,column=2).value==None:
        continue
    if sheet.cell(row=e_row,column=8).value==None:
        sheet.cell(row=e_row, column=8).value=0
    else:
        if sheet.cell(row=e_row, column=8).value<0:
            sheet.cell(row=e_row, column=8).value=-sheet.cell(row=e_row, column=8).value
    #----------------------全国
    if sheet.cell(row=e_row, column=2).value == '国家级' or sheet.cell(row=e_row, column=2).value == '省级':
        #去掉港澳台
        if sheet.cell(row=e_row, column=3).value == '澳门' or sheet.cell(row=e_row, column=3).value == '台湾' or sheet.cell(row=e_row, column=3).value == '香港':
            mm=1
        else:
            #print(sheet.cell(row=e_row, column=3).value)
            t = sheet.cell(row=e_row, column=1).value
            #print(sheet.cell(row=e_row, column=1).value)
            e_time = str(t.month) + "月" + str(t.day) + "日"
            if e_time not in country_list:
                country_list[e_time]={}
                country_list[e_time]['省级总和'] = [0, 0, 0]
            if sheet.cell(row=e_row, column=2).value == '国家级':
                country_list[e_time]['国家级']=[0,0,0]
                if sheet.cell(row=e_row, column=5).value != None:
                    country_list[e_time]['国家级'][0] = sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
                if sheet.cell(row=e_row, column=6).value != None:
                    country_list[e_time]['国家级'][1] = sheet.cell(row=e_row, column=6).value
                if sheet.cell(row=e_row, column=7).value != None:
                    country_list[e_time]['国家级'][2] = sheet.cell(row=e_row, column=7).value
            elif sheet.cell(row=e_row, column=2).value == '省级':
                country_list[e_time][sheet.cell(row=e_row, column=3).value]=[0,0,0]
                if sheet.cell(row=e_row, column=5).value != None:
                    country_list[e_time][sheet.cell(row=e_row, column=3).value][0] = sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
                if sheet.cell(row=e_row, column=6).value != None:
                    #print(sheet.cell(row=e_row, column=6).value)
                    country_list[e_time][sheet.cell(row=e_row, column=3).value][1] = sheet.cell(row=e_row, column=6).value
                if sheet.cell(row=e_row, column=7).value != None:
                    country_list[e_time][sheet.cell(row=e_row, column=3).value][2] = sheet.cell(row=e_row, column=7).value

                if sheet.cell(row=e_row, column=5).value != None:
                    country_list[e_time]['省级总和'][0] += sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
                if sheet.cell(row=e_row, column=6).value != None:
                    #print(sheet.cell(row=e_row, column=6).value)
                    country_list[e_time]['省级总和'][1] += sheet.cell(row=e_row, column=6).value
                if sheet.cell(row=e_row, column=7).value != None:
                    country_list[e_time]['省级总和'][2] += sheet.cell(row=e_row, column=7).value
        # if e_row > 199:
        #     print(e_row)
        #     print(country_list['1月31日']['省级总和'])
    if sheet.cell(row=e_row,column=2).value=='地区级' or sheet.cell(row=e_row,column=2).value=='省级':
        #---------------------------省级复核，每天---------------------------------
        #print(sheet.cell(row=e_row, column=2).value)
        #print(sheet.cell(row=e_row,column=2).value)
        #print(e_row)
        t=sheet.cell(row=e_row,column=1).value
        e_time = str(t.month) + "月" + str(t.day) + "日"
        if e_time not in time_list:
            time_list[e_time]={}
        if sheet.cell(row=e_row, column=3).value not in time_list[e_time]:
            time_list[e_time][sheet.cell(row=e_row, column=3).value] = [0, 0, 0, 0, 0, 0]

        if sheet.cell(row=e_row, column=2).value == '地区级':
            if sheet.cell(row=e_row, column=5).value != None:
                time_list[e_time][sheet.cell(row=e_row, column=3).value][0] += sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
            if sheet.cell(row=e_row, column=6).value != None:
                #print(sheet.cell(row=e_row, column=6).value)
                time_list[e_time][sheet.cell(row=e_row, column=3).value][1] += sheet.cell(row=e_row, column=6).value
            if sheet.cell(row=e_row, column=7).value != None:
                time_list[e_time][sheet.cell(row=e_row, column=3).value][2] += sheet.cell(row=e_row, column=7).value
        elif sheet.cell(row=e_row, column=2).value == '省级':
            if sheet.cell(row=e_row, column=5).value != None:
                time_list[e_time][sheet.cell(row=e_row, column=3).value][3] = sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
            if sheet.cell(row=e_row, column=6).value != None:
                time_list[e_time][sheet.cell(row=e_row, column=3).value][4] = sheet.cell(row=e_row, column=6).value
            if sheet.cell(row=e_row, column=7).value != None:
                time_list[e_time][sheet.cell(row=e_row, column=3).value][5] = sheet.cell(row=e_row, column=7).value
        # if e_row > 1456:
        #     print(e_row)
        #     print(time_list['1月23日']['海南'])
        #---------------------省级复核，累积-----------------------
        if sheet.cell(row=e_row, column=3).value not in postion_list:
            postion_list[sheet.cell(row=e_row, column=3).value]=[0,0,0,0,0,0]
        if sheet.cell(row=e_row, column=2).value=='地区级':
            if sheet.cell(row=e_row, column=5).value != None:
                postion_list[sheet.cell(row=e_row, column=3).value][0]+=sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
            if sheet.cell(row=e_row, column=6).value != None:
                postion_list[sheet.cell(row=e_row, column=3).value][1] += sheet.cell(row=e_row, column=6).value
            if sheet.cell(row=e_row, column=7).value != None:
                postion_list[sheet.cell(row=e_row, column=3).value][2] += sheet.cell(row=e_row, column=7).value
        elif sheet.cell(row=e_row, column=2).value == '省级':
            if sheet.cell(row=e_row, column=5).value != None:
                postion_list[sheet.cell(row=e_row, column=3).value][3] += sheet.cell(row=e_row, column=5).value-sheet.cell(row=e_row, column=8).value
            if sheet.cell(row=e_row, column=6).value != None:
                postion_list[sheet.cell(row=e_row, column=3).value][4] += sheet.cell(row=e_row, column=6).value
            if sheet.cell(row=e_row, column=7).value != None:
                postion_list[sheet.cell(row=e_row, column=3).value][5] += sheet.cell(row=e_row, column=7).value
# print(time_list['1月23日']['海南'])
# print(postion_list)
# print(country_list)
#------每天check
print("------------------省级复核，每天-----------------")
time_keys=list(time_list.keys())
for i in range(0,len(time_list)):
        time_k_keys=list(time_list[time_keys[i]].keys())
        for j in range(0,len(time_list[time_keys[i]])):
            a=time_list[time_keys[i]][time_k_keys[j]]
            #print(j,time_k_keys)
            if time_k_keys[j]==None:
                time_k_keys[j]='无数据'
            if a[0]!=a[3]:
                error_inf=time_k_keys[j]+','+time_keys[i]+',地区新增确诊病例累加值为'+str(int(a[0]))+',省级新增'+str(int(a[3]))
                print(error_inf)
            if a[1]!=a[4]:
                error_inf = time_k_keys[j] + ',' + time_keys[i] + ',地区新增治愈出院数累加值为' + str(int(a[1])) + ',省级新增' + str(int(a[4]))
                print(error_inf)
            if a[2] != a[5]:
                error_inf = time_k_keys[j] + ',' + time_keys[i] + ',地区新增死亡数累加值为' + str(int(a[2])) + ',省级新增' + str(int(a[5]))
                print(error_inf)
#---------累积check
print("------------------省级复核，累积-----------------")
postion_keys=list(postion_list.keys())
for i in range(0,len(postion_list)):
    a=postion_list[postion_keys[i]]
    if postion_keys[i] == None:
        postion_keys[i]='无数据'
    if a[0]!=a[3]:
        error_inf=postion_keys[i] + ',各地区新增确诊病例累积'+str(int(a[0]))+',省级新增'+str(int(a[3]))
        print(error_inf)
    if a[1]!=a[4]:
        error_inf = postion_keys[i] + ',各地区新增治愈出院数累积' + str(int(a[1])) + ',省级新增' + str(int(a[4]))
        print(error_inf)
    if a[2] != a[5]:
        error_inf = postion_keys[i] + ',各地区新增死亡数累积' + str(int(a[2])) + ',省级新增' + str(int(a[5]))
        print(error_inf)
#----------国家check
print("------------------全国级复核-----------------")
country_keys=list(country_list.keys())
for i in range(0,len(country_list)):
    a=country_list[country_keys[i]]
    #print(country_keys[i])
    if '国家级' not in list(a.keys()):
        error_inf = country_keys[i] + '没有统计国家数据'
        print(error_inf)
        continue
    if a['国家级'][0]!=a['省级总和'][0]:
        error_inf = country_keys[i] + ',各省新增确诊病例累积' + str(int(a['省级总和'][0])) + ',国家级新增' + str(int(a['国家级'][0]))
        print(error_inf)
    if a['国家级'][1]!=a['省级总和'][1]:
        error_inf = country_keys[i] + ',各省新增治愈出院数累积' + str(int(a['省级总和'][1])) + ',国家级新增' + str(int(a['国家级'][1]))
        print(error_inf)
    if a['国家级'][2]!=a['省级总和'][2]:
        error_inf = country_keys[i] + ',各省新增死亡数累积' + str(int(a['省级总和'][2])) + ',国家级新增' + str(int(a['国家级'][2]))
        print(error_inf)

