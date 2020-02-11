import openpyxl
from openpyxl.styles import Font, colors
from openpyxl.styles import  PatternFill
#获取excle文件
wb1 = openpyxl.load_workbook('data.xlsx')
wb2 = openpyxl.load_workbook('data2.xlsx')
#获取表格name
sheets1 = wb1.sheetnames
sheets2 = wb2.sheetnames
#准备
old_position=8
new_position=17
old_sheet=wb1[sheets1[old_position]]#省市数据
new_sheet=wb2['四川_0206']#统计数据
hight=0#编辑位置
#font = Font(color=colors.RED)#单元格样式
fill = PatternFill("solid", fgColor="FF0000")#单元格填充
new_hight=0
#-----表二对比-------
old_x=0
old_y=4
new_x=0
new_y=new_hight
old_list={}
i=0
old_sum=0
while True:
    if old_sheet.cell(row=old_y + 2 + i, column=old_x + 1).value==None:
        old_sum=old_sheet.cell(row=old_y + 2 + i, column=old_x + 2).value
        break
    old_list[old_sheet.cell(row=old_y + 2 + i, column=old_x + 1).value]=old_sheet.cell(row=old_y + 2 + i, column=old_x + 2).value
    i=i+1

j=0
new_sum=0
while True:
    #print(new_sheet.cell(row=new_y + 2 + j, column=new_x + 1).value)
    if new_sheet.cell(row=new_y + 2 + j, column=new_x + 1).value==None:
        new_sum=new_sheet.cell(row=new_y + 2 + j, column=new_x + 2).value
        break
    # new_list[j]=[new_sheet.cell(row=new_y + 2 + j, column=new_x + 1).value,new_sheet.cell(row=new_y + 2 + j, column=new_x + 2).value]
    else:
        if new_sheet.cell(row=new_y + 2 + j, column=new_x + 1).value in old_list:
            if new_sheet.cell(row=new_y + 2 + j, column=new_x + 2).value != old_list[new_sheet.cell(row=new_y + 2 + j, column=new_x + 1).value]:
                new_sheet.cell(row=new_y + 2 + j, column=new_x + 2).fill=fill
        else:
            new_sheet.cell(row=new_y + 2 + j, column=new_x + 1).fill = fill
            new_sheet.cell(row=new_y + 2 + j, column=new_x + 2).fill = fill
    j=j+1

if old_sum!=new_sum:
    new_sheet.cell(row=new_y + 2 + j, column=new_x + 2).fill=fill
new_hight=new_hight+new_y + 2 + j+2
#----表三-----
old_x=0
old_y=29
new_x=0
new_y=new_hight
for i in range(0,11):
    for j in range(0,3):
        if new_sheet.cell(row=new_y + 1 + i, column=new_x + j+2).value!=old_sheet.cell(row=old_y + 1 + i, column=old_x + j+2).value:
            new_sheet.cell(row=new_y + 1 + i, column=new_x + j+2).fill=fill
new_hight+=12
print(new_hight)
#----表四-----
old_x=0
old_y=47
new_x=0
new_y=new_hight
old_list.clear()
i=0
while True:
    print(old_sheet.cell(row=old_y + 1, column=old_x + 2 + i).value)
    if old_sheet.cell(row=old_y + 1, column=old_x + 2 + i).value==None:
        break
    t = old_sheet.cell(row=old_y + 1, column=old_x + 2 + i).value
    # print(t)
    # e_time=str(t.year)+"/"+str(t.month)+"/"+str(t.day)
    e_time = str(t.month) + "月" + str(t.day) + "日"
    old_list[e_time] = old_sheet.cell(row=old_y + 2,column=old_x + 2 + i).value
    i+=1
i=0
while True:
    print(new_sheet.cell(row=new_y + 1, column=new_x + 2 + i).value)
    if new_sheet.cell(row=new_y + 1, column=new_x + 2 + i).value==None:
        break
    if new_sheet.cell(row=new_y + 1, column=new_x + 2 + i).value in old_list:
        if new_sheet.cell(row=new_y + 2, column=new_x + 2 + i).value!=old_sheet.cell(row=old_y + 2, column=old_x + 2 + i).value:
            new_sheet.cell(row=new_y + 2, column=new_x + 2 + i).fill=fill
    else:
        new_sheet.cell(row=new_y + 1, column=new_x + 2 + i).fill=fill
        new_sheet.cell(row=new_y + 2, column=new_x + 2 + i).fill = fill
    i+=1
#保存（可以覆盖）
wb2.save('data2.xlsx')