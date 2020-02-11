import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
#获取excle文件
wb = openpyxl.load_workbook('data.xlsx')
#获取表格
sheets = wb.sheetnames
#目标文件

#print(sheets)
# for i in range(len(sheets)):
#     sheet = wb[sheets[i]]
#     #输出sheet名字
#     print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
#
#     for r in range(1, sheet.max_row + 1):
#         if r == 1:
#             print('\n' + ''.join(
#                 [str(sheet.cell(row=x+r, column=y+c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
#         else:
#             print(''.join([str(sheet.cell(row=x+r, column=y+c).value).ljust(20) for c in range(1, sheet.max_column + 1)]))
#     break

#----------------------工作区-------------------
#准备
data_position=3
#statistic_position=17
sheet=wb[sheets[data_position]]#省市数据
sheet2=wb.create_sheet("四川_0206") #创建工作表
hight=0#编辑位置
#---------表二---------
e2_x=0
e2_y=0#表二相对位置
sheet2.cell(row=e2_x + 1, column=e2_y + 1).value='地区'
sheet2.cell(row=e2_x + 1, column=e2_y + 2).value='数量'
sheet2.cell(row=e2_x + 1, column=e2_y + 3).value='确诊(重症)'
sheet2.cell(row=e2_x + 1, column=e2_y + 4).value='确诊(轻症)'
region_list=[]
region_value={}
list_len=0
sum=0
for r in range(2,sheet.max_row+1):
    if sheet.cell(row=r, column=1).value == None:
        continue
    sum+=1;
    if sheet.cell(row=r, column=20).value in region_list:
        region_value[sheet.cell(row=r, column=20).value]+=1
    else:
        list_len+=1;
        region_list.append(sheet.cell(row=r, column=20).value)
        region_value[sheet.cell(row=r, column=20).value]=1
#print(sum)
for r in range(0,list_len):#注意从0开始
    #print(region_list[r]);
    sheet2.cell(row=e2_x + 1 + (r+1), column=e2_y + 1).value=region_list[r]
    sheet2.cell(row=e2_x + 1 + (r+1), column=e2_y + 2).value=region_value[region_list[r]]
sheet2.cell(row=e2_x + 1 + list_len + 1, column=e2_y + 2).value=sum
hight=list_len+3#重置高度
#------------表三-----------
e3_x=hight
e3_y=0#表三相对位置
# print(sheet.title)
# print(sheet2.title)
#数据区
not_know=0
sum=0;
#代码区
sheet2.cell(row=e3_x + 1, column=e3_y + 1).value= '年龄'
sheet2.cell(row=e3_x + 1, column=e3_y + 2).value= '男性数量'
sheet2.cell(row=e3_x + 1, column=e3_y + 3).value= '女性数量'
sheet2.cell(row=e3_x + 1, column=e3_y + 4).value= '总人数'
sheet2.cell(row=e3_x + 2, column=e3_y + 1).value= '>70'
sheet2.cell(row=e3_x + 3, column=e3_y + 1).value= '61-70'
sheet2.cell(row=e3_x + 4, column=e3_y + 1).value= '51-60'
sheet2.cell(row=e3_x + 5, column=e3_y + 1).value= '41-50'
sheet2.cell(row=e3_x + 6, column=e3_y + 1).value= '31-40'
sheet2.cell(row=e3_x + 7, column=e3_y + 1).value= '21-30'
sheet2.cell(row=e3_x + 8, column=e3_y + 1).value= '11 至20'
sheet2.cell(row=e3_x + 9, column=e3_y + 1).value= '0-10'
sheet2.cell(row=e3_x + 10, column=e3_y + 1).value= '总'
sheet2.cell(row=e3_x + 11, column=e3_y + 1).value= '不明'
for i in range(2,10):
    for j in range(2,5):
        sheet2.cell(row=e3_x + i, column=e3_y + j).value=0

for r in range(2,sheet.max_row+1):
    #print(sheet.cell(row=x+r, column=y+1).value)
    if sheet.cell(row=r, column=1).value==None:
        continue
    sum+=1
    age=sheet.cell(row=r, column=5).value
    sex=sheet.cell(row=r, column=6).value
    #print(age,sex)
    if age==None or sex==None :
        not_know+=1
        #print(age, sex)
    elif age=='不明' or sex=='不明':
        not_know += 1
        #print(age, sex)
    elif int(age)>70:
        if sex=='男':
            sheet2.cell(row=e3_x + 2, column=e3_y + 2).value+=1
            #print(age, sex)
            #print(sheet2.cell(row=x+2,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 2, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+2, column=y+3).value)
    elif int(age)>60:
        if sex=='男':
            sheet2.cell(row=e3_x + 3, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+3,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 3, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+3, column=y+3).value)
    elif int(age)>50:
        if sex=='男':
            sheet2.cell(row=e3_x + 4, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+4,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 4, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+4, column=y+3).value)
    elif int(age)>40:
        if sex=='男':
            sheet2.cell(row=e3_x + 5, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+5,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 5, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+5, column=y+3).value)
    elif int(age)>30:
        if sex=='男':
            print(sheet2.cell(row=e3_x + 6, column=e3_y + 2).value)
            sheet2.cell(row=e3_x + 6, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+6,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 6, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+6, column=y+3).value)
    elif int(age)>20:
        if sex=='男':
            sheet2.cell(row=e3_x + 7, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+7,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 7, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+7, column=y+3).value)
    elif int(age)>10:
        if sex=='男':
            sheet2.cell(row=e3_x + 8, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+8,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 8, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+8, column=y+3).value)
    else:
        if sex=='男':
            sheet2.cell(row=e3_x + 9, column=e3_y + 2).value+=1
            # print(age, sex)
            # print(sheet2.cell(row=x+9,column=y+2).value)
        elif sex=='女':
            sheet2.cell(row=e3_x + 9, column=e3_y + 3).value += 1
            # print(age, sex)
            # print(sheet2.cell(row=x+8, column=y+3).value)
sheet2.cell(row=e3_x + 11, column=e3_y + 4).value = not_know#不明总数
sheet2.cell(row=e3_x + 12, column=e3_y + 4).value = sum#总人数
#print(sum)
#不同年龄总人数
for row2 in range(2,10):
    sheet2.cell(row=e3_x + row2, column=e3_y + 4).value= sheet2.cell(row=e3_x + row2, column=e3_y + 2).value + sheet2.cell(row=e3_x + row2, column=e3_y + 3).value
str1=str(get_column_letter(e3_y+2))+str(e3_x+2)
str2=str(get_column_letter(e3_y+2))+str(e3_x+9)
e_str="=SUM("+str1+":"+str2+")"
sheet2.cell(row=e3_x + 10, column=e3_y + 2).value= e_str
str1=str(get_column_letter(e3_y+3))+str(e3_x+2)
str2=str(get_column_letter(e3_y+3))+str(e3_x+9)
e_str="=SUM("+str1+":"+str2+")"
sheet2.cell(row=e3_x + 10, column=e3_y + 3).value= e_str
str1=str(get_column_letter(e3_y+4))+str(e3_x+2)
str2=str(get_column_letter(e3_y+4))+str(e3_x+9)
e_str="=SUM("+str1+":"+str2+")"
sheet2.cell(row=e3_x + 10, column=e3_y + 4).value= e_str
hight=hight+12
#----------表四--------
e4_x=hight+1
e4_y=0#表四的相对位置
sheet2.cell(row=e4_x + 1, column=e4_y + 1).value='时间'
sheet2.cell(row=e4_x + 2, column=e4_y + 1).value='每日新增确诊数量'
time_list=[]
time_value={}
list_len=0
for r in range(2,sheet.max_row+1):
    if sheet.cell(row=r, column=29).value == None or sheet.cell(row=r, column=29).value =='不明':
        continue
    t=sheet.cell(row=r, column=29).value
    #print(t)
    #e_time=str(t.year)+"/"+str(t.month)+"/"+str(t.day)
    e_time=str(t.month)+"月"+str(t.day)+"日"
    #print(e_time)
    if e_time in time_list:
        time_value[e_time]+=1
    else:
        list_len+=1;
        time_list.append(e_time)
        time_value[e_time]=1
#print(time_value)
for i in range(0,list_len):
    sheet2.cell(row=e4_x + 1, column=e4_y + 1+(i+1)).value=time_list[i]
    sheet2.cell(row=e4_x + 2, column=e4_y + 1+(i+1)).value=time_value[time_list[i]]
#保存（可以覆盖）
wb.save('data2.xlsx')
